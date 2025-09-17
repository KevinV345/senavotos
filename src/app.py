from multiprocessing import connection
import os
import random
from django import db
from flask import Flask, Response, jsonify, redirect, render_template, render_template_string, request, session, url_for, flash, make_response, send_file
from flask_mysqldb import MySQL
import MySQLdb.cursors
from functools import wraps
from base64 import b64encode
from markupsafe import Markup
from openpyxl import Workbook, load_workbook
import pandas as pd
import io
from datetime import datetime
import pymysql
from werkzeug.utils import secure_filename
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
import bcrypt
from datetime import datetime
import pytz
import bcrypt
import hashlib  

app = Flask(__name__)
app.secret_key = 'd4i8e2g1o7n#'

# Añadir el filtro en Jinja2 si no está habilitado:
@app.template_filter('b64encode')
def base64_encode(data):
    return b64encode(data).decode('utf-8')

# Configuración de la base de datos
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'senavotos'
# Configuración de la carpeta para cargar archivos
app.config['UPLOAD_FOLDER'] = 'uploads/'

# Verificar si la carpeta de carga existe y crearla si no
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}

mysql = MySQL(app)

# Decorador para evitar el almacenamiento en caché
def no_cache(view):
    @wraps(view)
    def cache_control(*args, **kwargs):
        response = make_response(view(*args, **kwargs))
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    return cache_control

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

@app.route('/', methods=['GET', 'POST'])
@no_cache
def home():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    mensaje = None

    # Obtener usuario desde la sesión
    usuario_sesion = session.get('usuario')

    if request.method == 'POST':
        documento = request.form.get('documento')
        # clave = request.form.get('clave')  # <<< CONTRASEÑA DESHABILITADA PARA PRUEBAS

        # Buscar usuario en la base de datos (ya no usamos la clave para autenticación)
        cursor.execute("""
            SELECT u.idusuario, u.documento, u.nombre, u.rol, u.jornada, u.asistencia_voto, f.clave
            FROM usuarios u 
            JOIN fichas f ON u.fichas_idfichas = f.idfichas
            WHERE u.documento = %s
        """, (documento,))
        usuario = cursor.fetchone()

        if usuario:
            # --- Bloque de contraseña comentado para pruebas ---
            # if usuario['clave'] == clave:
            #     ...
            # else:
            #     mensaje = "La clave ingresada es incorrecta. Por favor, inténtelo de nuevo."
            # ----------------------------------------------------

            # En vez de verificar clave, iniciamos sesión directamente si el documento existe
            # Validar si debe pasar por recepción
            if usuario['rol'] not in [2, 3, 4] and usuario['jornada'] != 'virtual' and usuario['asistencia_voto'] == 0:
                mensaje = "Debes pasar por recepción antes de continuar a la votación."
            else:
                session.permanent = True
                session['usuario'] = {
                    'documento': usuario['documento'],
                    'nombre': usuario['nombre'],
                    'rol': usuario['rol'],
                    'jornada': usuario['jornada'],
                    'idusuario': usuario['idusuario']
                }

                # Redireccionar según el rol con pantalla de carga
                if usuario['rol'] == 1:
                    destino = 'eleccion'
                elif usuario['rol'] == 2:
                    destino = 'admin'
                elif usuario['rol'] == 3:
                    destino = 'recepcionista'
                elif usuario['rol'] == 4:
                    destino = 'resultados'
                else:
                    destino = 'eleccion'

                return redirect(url_for('carga', destino=destino))

        else:
            mensaje = "El número de documento ingresado no está registrado. Por favor, inténtelo de nuevo."

    return render_template('index.html', mensaje=mensaje)


@app.route('/carga')
@no_cache
def carga():
    destino = request.args.get('destino', 'eleccion')
    return render_template('carga.html', destino=destino)

@app.route('/eleccion', methods=['GET', 'POST'])
@no_cache
def eleccion():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    mensaje = None

    # Obtener la hora actual en la zona horaria de Colombia
    zona_colombia = pytz.timezone("America/Bogota")
    hora_actual = datetime.now(zona_colombia).time()

    # Definir la hora de cierre (22:00 = 10:00 p.m.)
    hora_cierre = datetime.strptime("22:00", "%H:%M").time()

    # Si la hora actual es mayor o igual a la hora de cierre, redirigir a la página de elecciones cerradas
    if hora_actual >= hora_cierre:
        return render_template('eleccion_cerrada.html')

    # Validar sesión
    usuario_sesion = session.get('usuario')
    if not usuario_sesion:
        return redirect(url_for('home'))

    documento = usuario_sesion['documento']
    idusuario = usuario_sesion['idusuario']

    # Verificar si el usuario ya votó
    cursor.execute("SELECT * FROM votos WHERE usuarios_idusuario = %s", (idusuario,))
    voto_existente = cursor.fetchone()
    if voto_existente:
        mensaje = "Ya has votado. No puedes realizar más de un voto."
        return render_template('eleccion.html', candidatos=[], mensaje=mensaje)

    # Obtener candidatos
    jornada_usuario = usuario_sesion['jornada']
    cursor.execute(
        "SELECT idcandidato, nombre_candidato, foto FROM candidatos WHERE jornada = %s",
        (jornada_usuario,)
    )
    candidatos = cursor.fetchall()

    # Convertir fotos a base64
    for candidato in candidatos:
        if candidato['foto']:
            candidato['foto'] = b64encode(candidato['foto']).decode('utf-8')

    if request.method == 'POST':
        candidato_id = request.form.get('candidato_id')
        if candidato_id:
            cursor.execute(
                "INSERT INTO votos (usuarios_idusuario, candidatos_idcandidato, fecha_hora) VALUES (%s, %s, NOW())",
                (idusuario, candidato_id)
            )
            mysql.connection.commit()
            return redirect(url_for('gracias'))
        else:
            mensaje = "Debe seleccionar un candidato para continuar."

    return render_template('eleccion.html', candidatos=candidatos, mensaje=mensaje)

@app.route('/gracias')
@no_cache
def gracias():
    # Renderiza una página temporal de agradecimiento
    return render_template('gracias.html')

@app.route('/admin')
@no_cache
def admin():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Obtener los votos y unir con datos de usuarios, candidatos y fichas
    try:
        cursor.execute(""" 
            SELECT 
                u.documento, 
                u.nombre AS votante, 
                c.nombre_candidato, 
                u.jornada, 
                f.idfichas AS ficha
            FROM votos v
            JOIN usuarios u ON v.usuarios_idusuario = u.idusuario
            JOIN candidatos c ON v.candidatos_idcandidato = c.idcandidato
            JOIN fichas f ON u.fichas_idfichas = f.idfichas
        """)
        votos = cursor.fetchall()
    except MySQLdb.Error as e:
        # print(f"Error al consultar los votos: {e}")
        votos = []

    return render_template('admin.html', votos=votos)

@app.route('/admin/votos')
def actualizar_votos():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    try:
        cursor.execute(""" 
            SELECT 
                u.documento, 
                u.nombre AS votante, 
                c.nombre_candidato, 
                u.jornada, 
                f.idfichas AS ficha
            FROM votos v
            JOIN usuarios u ON v.usuarios_idusuario = u.idusuario
            JOIN candidatos c ON v.candidatos_idcandidato = c.idcandidato
            JOIN fichas f ON u.fichas_idfichas = f.idfichas
        """)
        votos = cursor.fetchall()
    except MySQLdb.Error:
        votos = []

    jornadas = sorted(set(v['jornada'] for v in votos))
    candidatos = sorted(set(v['nombre_candidato'] for v in votos))

    html_filas = render_template_string("""
        {% for voto in votos %}
        <tr>
            <td>{{ voto.documento }}</td>
            <td>{{ voto.votante }}</td>
            <td>{{ voto.ficha or 'Sin ficha' }}</td>
            <td>{{ voto.jornada }}</td>
            <td>{{ voto.nombre_candidato }}</td>
        </tr>
        {% endfor %}
    """, votos=votos)

    return jsonify({
        "html": html_filas,
        "jornadas": jornadas,
        "candidatos": candidatos,
        "total": len(votos)  # Total de votos
    })

@app.route('/admin/exportar_excel', methods=['POST'])
def exportar_excel():
    jornada_filtro = request.form.get('jornada')
    candidato_filtro = request.form.get('candidato')

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    sql = """ 
        SELECT 
            u.documento, 
            u.nombre AS votante,  
            u.jornada, 
            COALESCE(f.idfichas, 'Sin ficha') AS ficha,
            c.nombre_candidato
        FROM votos v
        JOIN usuarios u ON v.usuarios_idusuario = u.idusuario
        JOIN candidatos c ON v.candidatos_idcandidato = c.idcandidato
        LEFT JOIN fichas f ON u.fichas_idfichas = f.idfichas
        WHERE 1=1
    """
    filtros = []

    if jornada_filtro:
        sql += " AND u.jornada = %s"
        filtros.append(jornada_filtro)

    if candidato_filtro:
        sql += " AND c.nombre_candidato = %s"
        filtros.append(candidato_filtro)

    cursor.execute(sql, tuple(filtros))
    votos = cursor.fetchall()

    wb = Workbook()

    # -------------------- Hoja Votos --------------------
    ws = wb.active
    ws.title = "Votos"

    encabezados = ["Documento del Votante", "Nombre del Votante", "Ficha", "Jornada"]
    header_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    ws.append(encabezados)
    for col_num, header in enumerate(encabezados, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    for voto in votos:
        ws.append([voto['documento'], voto['votante'], voto['ficha'], voto['jornada']])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # -------------------- Hoja Resumen --------------------
    ws2 = wb.create_sheet("Resumen")

    # ==== Tabla 1: Votos por Jornada ====
    ws2.append(["Jornada", "Total Votos"])
    for col in range(1, 3):
        cell = ws2.cell(row=1, column=col)
        cell.fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    colores_jornadas = {
        "mañana": "FFD700",   # Dorado
        "tarde": "87CEEB",    # Azul cielo
        "virtual": "90EE90",  # Verde claro
        "mixta": "FFB6C1"     # Rosado
    }

    jornadas = {"mañana": 0, "tarde": 0, "virtual": 0, "mixta": 0}
    for voto in votos:
        jornadas[voto['jornada']] += 1

    row = 2
    for jornada, total in jornadas.items():
        nombre_capitalizado = jornada.capitalize()
        ws2.append([nombre_capitalizado, total])
        for col in range(1, 3):
            cell = ws2.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 1:
                cell.fill = PatternFill(start_color=colores_jornadas[jornada], 
                                        end_color=colores_jornadas[jornada], 
                                        fill_type="solid")
        row += 1

    # ==== Tabla 2: Votos por Candidato ====
    ws2.append([])
    ws2.append(["Candidato", "Total Votos"])
    for col in range(1, 3):
        cell = ws2.cell(row=row+1, column=col)
        cell.fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    candidatos_count = {}
    for voto in votos:
        candidato = voto['nombre_candidato']
        candidatos_count[candidato] = candidatos_count.get(candidato, 0) + 1

    row += 2
    for candidato, total in candidatos_count.items():
        ws2.append([candidato, total])
        for col in range(1, 3):
            cell = ws2.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    # ==== Ajustar anchos ====
    for col in range(1, 3):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    # ==== Gráfico de Barras ====
    chart = BarChart()
    chart.type = "col"
    chart.title = "Votos por Jornada"
    chart.y_axis.title = "Votos"
    chart.x_axis.title = "Jornada"

    # Cada jornada como barra con su color
    col = 2  # columna "Total Votos"
    for i, (jornada, total) in enumerate(jornadas.items(), start=2):
        data = Reference(ws2, min_col=col, min_row=i, max_row=i)  # solo esa fila
        series = Series(data, title=ws2.cell(row=i, column=1).value)  # título = nombre jornada
        series.graphicalProperties.solidFill = list(colores_jornadas.values())[i-2]  # color por jornada
        series.data_labels = DataLabelList()
        series.data_labels.showVal = True
        chart.append(series)

    # Categorías (Mañana, Tarde, etc.)
    categories = Reference(ws2, min_col=1, min_row=2, max_row=5)
    chart.set_categories(categories)

    ws2.add_chart(chart, "E2")

    # -------------------- Guardar y enviar --------------------
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, as_attachment=True, download_name="lista_votos.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
@app.route('/upload', methods=['GET', 'POST'])
@no_cache
def upload():
    return render_template('upload.html')

@app.route('/procesar_fichas', methods=['POST'])
@no_cache
def procesar_fichas():
    file = request.files.get('file')
    jornada = request.form.get('jornada')

    if not jornada or jornada not in ['mañana', 'tarde', 'mixta', 'virtual']:
        flash('Por favor, selecciona una jornada válida.', 'danger')
        return redirect(url_for('upload'))

    if not file or not file.filename.lower().endswith('.xlsx'):
        flash('Por favor, sube un archivo Excel válido (.xlsx).', 'danger')
        return redirect(url_for('upload'))

    try:
        df = pd.read_excel(file, sheet_name=0, header=None)

        ficha_combined = df.iloc[1, 2]
        if pd.isna(ficha_combined) or ficha_combined.strip() == '':
            raise ValueError("El campo 'Ficha de Caracterización' está vacío.")

        ficha_combined = ficha_combined.strip()
        if '-' in ficha_combined:
            id_ficha, nombre_programa = ficha_combined.split('-', 1)
            id_ficha = id_ficha.strip()
            nombre_programa = nombre_programa.strip()
        else:
            raise ValueError("Formato inválido: se esperaba 'ID - Nombre del programa'.")

        clave_original = f"{id_ficha[:3]}-{nombre_programa[:3]}".upper()
        clave_hash = hashlib.sha256(clave_original.encode('utf-8')).hexdigest()

        # Intentar insertar la ficha
        try:
            with mysql.connection.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO fichas (idfichas, nombre_programa, clave) VALUES (%s, %s, %s)",
                    (id_ficha, nombre_programa, clave_hash)
                )
            mysql.connection.commit()
        except pymysql.err.IntegrityError as e:
            if e.args[0] == 1062:
                flash(f"La ficha con ID {id_ficha} ya existe en el sistema.", "danger")
                return redirect(url_for('upload'))
            else:
                raise

        # Procesar aprendices (desde fila 6, columnas B-F)
        for row in df.iloc[5:33].itertuples(index=False):
            documento = row[1]
            nombre = row[2]
            apellidos = row[3]
            celular = str(row[4]).strip() if len(row) > 4 and not pd.isna(row[4]) else None
            correo = str(row[5]).strip() if len(row) > 5 and not pd.isna(row[5]) else None

            if pd.isna(documento) or pd.isna(nombre) or pd.isna(apellidos):
                continue

            nombre_completo = f"{nombre.strip()} {apellidos.strip()}"

            with mysql.connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO usuarios (documento, nombre, celular, correo, fichas_idfichas, jornada)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    (documento, nombre_completo, celular, correo, id_ficha, jornada)
                )
        mysql.connection.commit()

        flash("Datos de la ficha y aprendices guardados correctamente.", "success")

    except Exception as e:
        flash("Error al procesar el archivo. Verifica que el formato sea correcto.", "danger")
        print("[ERROR]", str(e))  # Para ti como desarrollador

    return redirect(url_for('upload'))

@app.route('/recepcionista')
@no_cache
def recepcionista():
    return render_template('recepcionista.html')

@app.route('/buscar_votante', methods=['POST'])
@no_cache
def buscar_votante():
    data = request.get_json()
    documento = data.get('documento')

    if not documento:
        return jsonify({"status": "error", "message": "Documento no proporcionado."})

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Modificación: se incluyen los campos 'ficha' y 'mesa'
    cursor.execute("""
        SELECT idusuario, nombre, asistencia_voto, mesa, fichas_idfichas AS ficha, jornada 
        FROM usuarios 
        WHERE documento = %s
    """, (documento,))
    usuario = cursor.fetchone()

    if usuario:
        if usuario['asistencia_voto'] == 0:
            # Lista global de asignaciones mezcladas
            if 'cola_mesas' not in session or not session['cola_mesas']:
                mesas = list(range(1, 11))  # [1..10]
                random.shuffle(mesas)
                session['cola_mesas'] = mesas

            # Sacar el siguiente computador de la cola
            mesa_asignada = session['cola_mesas'].pop(0)

            # Actualizar asistencia, ficha y número de mesa
            cursor.execute("""
                UPDATE usuarios 
                SET asistencia_voto = 1, mesa = %s 
                WHERE idusuario = %s
            """, (mesa_asignada, usuario['idusuario']))
            mysql.connection.commit()

            return jsonify({
                "status": "success",
                "message": f"{usuario['nombre']} Listo Para Votar.",
                "data": {
                    "documento": documento,
                    "nombre": usuario['nombre'],
                    "ficha": usuario['ficha'],
                    "jornada": usuario['jornada'],
                    "mesa": mesa_asignada
                }
            })
        else:
            return jsonify({
                "status": "warning",
                "message": f"{usuario['nombre']} Ya Votó.",
                "data": {
                    "documento": documento,
                    "nombre": usuario['nombre'],
                    "ficha": usuario['ficha'],
                    "jornada": usuario['jornada'],
                    "mesa": usuario['mesa']
                }
            })
    else:
        return jsonify({"status": "error", "message": "Documento no Registrado."})
    
@app.route('/crear_candidato', methods=['GET', 'POST'])
@no_cache
def crear_candidato():
    try:
        if request.method == 'POST':
            nombre_candidato = request.form['nombre_candidato']
            jornada = request.form['jornada']
            foto_candidato = request.files['foto_candidato']
            
            foto_binario = foto_candidato.read()

            cursor = mysql.connection.cursor()
            sql = "INSERT INTO candidatos (nombre_candidato, foto, jornada) VALUES (%s, %s, %s)"
            values = (nombre_candidato, foto_binario, jornada)
            cursor.execute(sql, values)
            mysql.connection.commit()

            flash('¡Candidato creado exitosamente!', 'success')
            return redirect(url_for('crear_candidato'))
    except Exception as e:
        flash('Hubo un error al crear el candidato. Inténtalo de nuevo.', 'error')

    # 🔹 Aquí cargamos la lista de candidatos agrupados por nombre
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("""
        SELECT 
            MIN(idcandidato) AS idcandidato, 
            nombre_candidato, 
            GROUP_CONCAT(jornada ORDER BY jornada SEPARATOR ', ') AS jornadas
        FROM candidatos
        GROUP BY nombre_candidato
    """)
    candidatos = cursor.fetchall()

    return render_template('crear_candidato.html', candidatos=candidatos)

# --- LISTAR CANDIDATOS ---
@app.route("/candidatos")
def candidatos():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("""
        SELECT 
            MIN(idcandidato) AS idcandidato, 
            nombre_candidato, 
            GROUP_CONCAT(jornada ORDER BY jornada SEPARATOR ', ') AS jornadas
        FROM candidatos
        GROUP BY nombre_candidato
    """)
    lista = cursor.fetchall()
    return render_template("crear_candidato.html", candidatos=lista)


# --- MOSTRAR FOTO ---
@app.route("/foto_candidato/<int:idcandidato>")
def foto_candidato(idcandidato):
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT foto FROM candidatos WHERE idcandidato = %s", (idcandidato,))
    row = cursor.fetchone()
    if row and row[0]:
        return Response(row[0], mimetype="image/jpeg")
    return "Sin foto", 404

# --- EDITAR CANDIDATO ---
@app.route("/editar_candidato/<int:idcandidato>", methods=["GET", "POST"])
def editar_candidato(idcandidato):
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    if request.method == "POST":
        nombre = request.form["nombre_candidato"]
        jornada = request.form["jornada"]

        if "foto_candidato" in request.files and request.files["foto_candidato"].filename != "":
            foto = request.files["foto_candidato"].read()
            cursor.execute(
                "UPDATE candidatos SET nombre_candidato=%s, jornada=%s, foto=%s WHERE idcandidato=%s",
                (nombre, jornada, foto, idcandidato),
            )
        else:
            cursor.execute(
                "UPDATE candidatos SET nombre_candidato=%s, jornada=%s WHERE idcandidato=%s",
                (nombre, jornada, idcandidato),
            )

        mysql.connection.commit()
        flash("Candidato actualizado correctamente", "success")
        return redirect(url_for("candidatos"))

    cursor.execute("SELECT * FROM candidatos WHERE idcandidato = %s", (idcandidato,))
    candidato = cursor.fetchone()
    return render_template("editar_candidato.html", candidato=candidato)

# --- ELIMINAR CANDIDATO ---
@app.route("/eliminar_candidato/<int:idcandidato>")
def eliminar_candidato(idcandidato):
    cursor = mysql.connection.cursor()
    cursor.execute("DELETE FROM candidatos WHERE idcandidato = %s", (idcandidato,))
    mysql.connection.commit()
    flash("Candidato eliminado correctamente", "success")
    return redirect(url_for("candidatos"))

@app.route('/resultados', methods=['GET'])
@no_cache
def resultados():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    try:
        cursor.execute("""
            SELECT 
                c.idcandidato, 
                c.nombre_candidato, 
                COUNT(v.idvoto) AS total_votos,
                c.jornada
            FROM 
                candidatos c 
            LEFT JOIN 
                votos v 
            ON 
                c.idcandidato = v.candidatos_idcandidato 
            GROUP BY 
                c.idcandidato, c.jornada
        """)
        resultados = cursor.fetchall()
    except MySQLdb.Error as e:
        # print(f"Error al consultar los resultados: {e}")
        resultados = []

    # Agrupar resultados por jornada
    resultados_por_jornada = {
        "mañana": [],
        "tarde": [],
        "mixta": [],
        "virtual": []
    }

    for resultado in resultados:
        jornada = resultado["jornada"]
        if jornada in resultados_por_jornada:
            resultados_por_jornada[jornada].append(resultado)

    return render_template('resultados.html', resultados=resultados_por_jornada)

@app.route('/resultados/datos', methods=['GET'])
def actualizar_resultados():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    try:
        cursor.execute("""
            SELECT 
                c.idcandidato, 
                c.nombre_candidato, 
                COUNT(v.idvoto) AS total_votos,
                c.jornada
            FROM 
                candidatos c 
            LEFT JOIN 
                votos v 
            ON 
                c.idcandidato = v.candidatos_idcandidato 
            GROUP BY 
                c.idcandidato, c.jornada
        """)
        resultados = cursor.fetchall()
    except MySQLdb.Error:
        resultados = []

    # Agrupar resultados por jornada
    resultados_por_jornada = {
        "mañana": [],
        "tarde": [],
        "mixta": [],
        "virtual": []
    }

    for resultado in resultados:
        jornada = resultado["jornada"]
        if jornada in resultados_por_jornada:
            resultados_por_jornada[jornada].append(resultado)

    return jsonify(resultados_por_jornada)

if __name__ == '__main__':
    app.run(debug=True)